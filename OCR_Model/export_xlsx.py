import statistics
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from structure import TableStructure

FLAG_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")


def _flagged_cells(struct, row_order):
    """
    Not a correction (there's no word list to correct against) -- a QA
    surface. A cell whose text is a drastic length-outlier for its own
    column (e.g. a 2-char read in a column where every other row is a
    long description) is the visible signature of a model hallucination
    like '55' standing in for 'PIPE ASSY TEE-3 TO TCV-42', or a stray CJK
    glyph like '一' standing in for '1'. Flags these (row, col) pairs so
    they render highlighted instead of passing silently as clean data.
    """
    by_col = {}
    for r in row_order:
        for c in range(struct.n_cols):
            cell = struct.cells[(r, c)]
            if cell.merge_flag:
                continue  # merge blocks are legitimately short/shared; not in scope here
            txt = (cell.ocr_text or "").strip()
            if txt:
                by_col.setdefault(c, []).append((r, len(txt)))

    flagged = set()
    for c, entries in by_col.items():
        lengths = [n for _, n in entries]
        if len(lengths) < 4:
            continue  # too few samples in this column to judge an outlier
        med = statistics.median(lengths)
        if med < 4:
            continue  # column itself normally holds short values (e.g. QTY) -- not comparable
        for r, n in entries:
            if n <= 3 and n < 0.4 * med:
                flagged.add((r, c))
    return flagged


def export(struct: TableStructure, out_path, col_names=None, row_order=None):
    """
    col_names: one name per column. No longer a fixed schema -- pass in
    whatever header.build_header() discovered for this sheet. Falls back
    to generic "Col1", "Col2", ... only if nothing was supplied.

    row_order: struct row indices to write, in the order they should
    appear (top to bottom) in the exported sheet. Defaults to the raw
    image row order (0..n_rows-1) for backward compatibility, but for a
    bottom-up table this should be header.build_header()'s
    data_row_order so rows come out in correct reading order instead of
    upside down.
    """
    col_names = col_names or [f"Col{i + 1}" for i in range(struct.n_cols)]
    row_order = row_order if row_order is not None else list(range(struct.n_rows))
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Table"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, name in enumerate(col_names):
        cell = ws.cell(row=1, column=i + 1, value=name)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Vertical dimension: NEVER merged in Excel. Every logical row stays its
    # own row; a value that spans multiple rows is simply repeated into
    # each row's own cell rather than spanned.
    #
    # Horizontal dimension: DOES get a real Excel merge. If a merge block
    # covers multiple columns, each individual row within that block gets
    # its own horizontal merge across those columns (written once into the
    # leftmost cell of that row's span) -- rows are still never merged
    # with each other, only columns within the same row.
    written_horizontal_spans = set()  # (row, col_start) already merged, avoid re-merging
    flagged = _flagged_cells(struct, row_order)

    for excel_row_offset, r in enumerate(row_order):
        excel_row = excel_row_offset + 2
        c = 0
        while c < struct.n_cols:
            cell = struct.cells[(r, c)]
            excel_col = c + 1

            if cell.merge_flag:
                region = next(m for m in struct.merge_regions if m.merge_id == cell.merge_id)
                col_span = region.col_end - region.col_start  # 0 if single column
                if col_span > 0:
                    # horizontal merge: merge THIS row's cells across the block's columns
                    start_col = region.col_start + 1
                    end_col = region.col_end + 1
                    key = (r, start_col)
                    xcell = ws.cell(row=excel_row, column=start_col, value=cell.ocr_text)
                    xcell.border = border
                    xcell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                    if key not in written_horizontal_spans:
                        ws.merge_cells(start_row=excel_row, start_column=start_col,
                                        end_row=excel_row, end_column=end_col)
                        written_horizontal_spans.add(key)
                    c = region.col_end + 1
                    continue
                else:
                    # pure vertical-only merge (single column, multiple rows):
                    # no Excel merge, just write this row's own (repeated) value
                    xcell = ws.cell(row=excel_row, column=excel_col, value=cell.ocr_text)
                    xcell.border = border
                    xcell.alignment = Alignment(vertical='center', wrap_text=True)
                    c += 1
                    continue
            else:
                xcell = ws.cell(row=excel_row, column=excel_col, value=cell.ocr_text)
                xcell.border = border
                xcell.alignment = Alignment(vertical='center', wrap_text=True)
                if (r, c) in flagged:
                    xcell.fill = FLAG_FILL
                c += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 16

    wb.save(out_path)
    return out_path


def export_structure_demo(struct: TableStructure, out_path, col_names=None):
    """
    Same layout/merge rules as export(), but instead of writing the (OCR-
    quality-dependent) recognized text, every merge block is filled with
    an explicit group label -- e.g. 'GROUP 4 (rows 9-13)'. This proves
    which rows/columns the structure stage believes share one value,
    completely decoupled from whether OCR read that value correctly.
    Standalone (non-merged) cells show their row/col identity instead of
    OCR text, for the same reason.
    """
    col_names = col_names or [f"Col{i}" for i in range(struct.n_cols)]
    wb = Workbook()
    ws = wb.active
    ws.title = "Structure Demo"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, name in enumerate(col_names):
        cell = ws.cell(row=1, column=i + 1, value=name)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    written_horizontal_spans = set()

    for r in range(struct.n_rows):
        excel_row = r + 2
        c = 0
        while c < struct.n_cols:
            cell = struct.cells[(r, c)]
            excel_col = c + 1

            if cell.merge_flag:
                region = next(m for m in struct.merge_regions if m.merge_id == cell.merge_id)
                col_span = region.col_end - region.col_start
                label = f"GROUP {region.merge_id} (rows {region.row_start+1}-{region.row_end+1})"
                if col_span > 0:
                    start_col = region.col_start + 1
                    end_col = region.col_end + 1
                    key = (r, start_col)
                    xcell = ws.cell(row=excel_row, column=start_col, value=label)
                    xcell.border = border
                    xcell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                    xcell.font = Font(color="0000FF")
                    if key not in written_horizontal_spans:
                        ws.merge_cells(start_row=excel_row, start_column=start_col,
                                        end_row=excel_row, end_column=end_col)
                        written_horizontal_spans.add(key)
                    c = region.col_end + 1
                    continue
                else:
                    xcell = ws.cell(row=excel_row, column=excel_col, value=label)
                    xcell.border = border
                    xcell.alignment = Alignment(vertical='center', wrap_text=True)
                    xcell.font = Font(color="0000FF")
                    c += 1
                    continue
            else:
                xcell = ws.cell(row=excel_row, column=excel_col, value="")
                xcell.border = border
                c += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 22

    wb.save(out_path)
    return out_path
