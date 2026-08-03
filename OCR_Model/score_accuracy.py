"""
Compares the pipeline's own extracted.xlsx output against a hand-corrected
ground-truth xlsx (e.g. extracted_corrected.xlsx), cell by cell.
"""
import sys
import argparse
from difflib import SequenceMatcher
import openpyxl


def load_grid(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    n_rows = ws.max_row
    n_cols = ws.max_column
    grid = [["" for _ in range(n_cols)] for _ in range(n_rows)]

    for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols):
        for cell in row:
            v = cell.value
            grid[cell.row - 1][cell.column - 1] = "" if v is None else str(v).strip()

    for merged_range in ws.merged_cells.ranges:
        top_val = grid[merged_range.min_row - 1][merged_range.min_col - 1]
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                grid[r - 1][c - 1] = top_val

    header = grid[0]
    data = grid[1:]
    return header, data


def normalize(s):
    return " ".join(s.split()).upper()


def similarity(a, b):
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


def score(pred_path, truth_path, fuzzy_threshold=0.9):
    pred_header, pred_rows = load_grid(pred_path)
    truth_header, truth_rows = load_grid(truth_path)

    print(f"Predicted file header ({len(pred_header)} cols): {pred_header}")
    print(f"Truth file header     ({len(truth_header)} cols): {truth_header}")

    n_rows = min(len(pred_rows), len(truth_rows))
    n_cols = min(len(pred_header), len(truth_header))
    if len(pred_rows) != len(truth_rows):
        print(f"\n WARNING: row count differs (predicted={len(pred_rows)}, "
              f"truth={len(truth_rows)}). Comparing only the first {n_rows} "
              f"rows -- check row order / header discovery if this is unexpected.\n")
    if len(pred_header) != len(truth_header):
        print(f"\n WARNING: column count differs (predicted={len(pred_header)}, "
              f"truth={len(truth_header)}). Comparing only the first {n_cols} "
              f"columns by POSITION, not by header name.\n")

    total = 0
    exact_matches = 0
    fuzzy_matches = 0
    per_col_total = [0] * n_cols
    per_col_exact = [0] * n_cols
    mismatches = []

    for r in range(n_rows):
        for c in range(n_cols):
            pred_val = pred_rows[r][c]
            truth_val = truth_rows[r][c]
            if not pred_val and not truth_val:
                continue

            total += 1
            per_col_total[c] += 1
            sim = similarity(pred_val, truth_val)

            if normalize(pred_val) == normalize(truth_val):
                exact_matches += 1
                per_col_exact[c] += 1
                fuzzy_matches += 1
            else:
                if sim >= fuzzy_threshold:
                    fuzzy_matches += 1
                mismatches.append((r + 2, c + 1, truth_header[c] if c < len(truth_header) else f"col{c+1}",
                                    truth_val, pred_val, sim))

    exact_acc = exact_matches / total if total else 0.0
    fuzzy_acc = fuzzy_matches / total if total else 0.0

    print(f"\n=== OVERALL ===")
    print(f"Cells compared: {total}")
    print(f"Exact-match accuracy: {exact_acc:.1%}  ({exact_matches}/{total})")
    print(f"Fuzzy-match accuracy (>= {fuzzy_threshold:.0%} similarity): "
          f"{fuzzy_acc:.1%}  ({fuzzy_matches}/{total})")

    print(f"\n=== PER-COLUMN EXACT-MATCH ACCURACY ===")
    for c in range(n_cols):
        name = truth_header[c] if c < len(truth_header) else f"col{c+1}"
        t = per_col_total[c]
        e = per_col_exact[c]
        acc = e / t if t else float("nan")
        print(f"  {name!r:30} {e:>3}/{t:<3}  ({acc:.0%})" if t else f"  {name!r:30} (no cells)")

    if mismatches:
        print(f"\n=== MISMATCHES ({len(mismatches)}) ===")
        mismatches.sort(key=lambda m: m[5])
        for excel_row, excel_col, col_name, truth_val, pred_val, sim in mismatches:
            print(f"  row {excel_row:>3} col {excel_col} [{col_name}]  "
                  f"sim={sim:.2f}\n"
                  f"    truth: {truth_val!r}\n"
                  f"    pred : {pred_val!r}")

    return exact_acc, fuzzy_acc


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("predicted_xlsx")
    ap.add_argument("truth_xlsx")
    ap.add_argument("--fuzzy-threshold", type=float, default=0.9)
    args = ap.parse_args()
    score(args.predicted_xlsx, args.truth_xlsx, fuzzy_threshold=args.fuzzy_threshold)


if __name__ == "__main__":
    main()
