"""
Harvest real, ground-truth-labeled LETTER glyphs from the PART DESCRIPTION
column, using the already-hand-verified text in extracted_corrected.xlsx as
ground truth (same trust model as harvest_digits.py's use of the SL column).

Approach: word-level segmentation by ink-gap WIDTH, not global pitch
calibration. A gap between words (a real space) is much wider than the
1-2px internal gaps this stencil font's individual letters have (see the
earlier "PIPE ASSY APU-23 TO PRV" over-segmentation test), so splitting
on wide gaps reliably recovers word boundaries. Each recovered word chunk
is then matched positionally against the known text's word list; if word
counts don't match for a row, that row is skipped (logged) rather than
guessed at.

Requires: extracted_corrected.xlsx and result.png in the working directory.

Usage:
    python3 harvest_letters.py
"""
import numpy as np
import openpyxl
from PIL import Image

IMAGE_PATH = "result.png"
XLSX_PATH = "extracted_corrected.xlsx"

# From the grid-line detection done earlier on this document.
ROW_BOUNDS = [73, 92, 111, 131, 150, 170, 189, 209, 228, 247, 267, 286, 306,
              325, 345, 364, 383, 403, 422, 442, 461, 481, 500, 519, 539,
              558, 578, 597, 617, 636, 655, 675, 694, 714, 733, 753, 772,
              791, 811, 830, 850, 869, 889, 908, 927, 947, 966, 986, 1005,
              1025, 1044]
DESC_X0, DESC_X1 = 47, 320  # inset from grid lines (44, 321) to dodge border bleed
INK_THRESHOLD = 150
WORD_GAP_MIN_PX = 5   # >=5 consecutive zero-ink columns => treat as a word
                       # boundary (space), not an intra-letter stencil gap
TARGET = 22


def tight_crop(bin_img):
    ys, xs = np.where(bin_img > 0)
    if len(xs) == 0:
        return None
    return bin_img[ys.min():ys.max() + 1, xs.min():xs.max() + 1]


def resize_glyph(glyph_bin):
    g = Image.fromarray((glyph_bin * 255).astype(np.uint8))
    g = g.resize((TARGET, TARGET), Image.LANCZOS)
    return np.array(g) / 255.0


def split_into_words(bw_row):
    """bw_row: 2D binary array (1=ink), already row-cropped to the cell
    height. Returns a list of (x0, x1) pixel spans, one per detected word,
    based on gap width rather than absolute pitch."""
    col_ink = bw_row.sum(axis=0)
    has_ink = col_ink > 1  # >1, not >0: filters the 1px horizontal grid-line
                            # bleed that affects every column uniformly
                            # (same fix as harvest_digits.py needed)
    words = []
    in_word = False
    start = 0
    gap_run = 0
    for i, v in enumerate(has_ink):
        if v:
            if not in_word:
                start = i
                in_word = True
            gap_run = 0
        else:
            if in_word:
                gap_run += 1
                if gap_run >= WORD_GAP_MIN_PX:
                    words.append((start, i - gap_run + 1))
                    in_word = False
    if in_word:
        words.append((start, len(has_ink)))
    return words


def slice_word_into_chars(word_bin, n_chars):
    """Equal-width slicing of a single word's tight-cropped ink region
    into n_chars pieces -- same technique validated on digits."""
    w = word_bin.shape[1]
    edges = np.linspace(0, w, n_chars + 1).astype(int)
    glyphs = []
    for i in range(n_chars):
        piece = word_bin[:, edges[i]:edges[i + 1]]
        pt = tight_crop(piece)
        if pt is None or pt.shape[0] < 3 or pt.shape[1] < 2:
            return None
        glyphs.append(pt)
    return glyphs


def main():
    im = Image.open(IMAGE_PATH).convert("L")
    arr = np.array(im)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    # Build SL -> PART DESCRIPTION ground truth from the corrected sheet.
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    sl_col = header.index("SL\nNO") if "SL\nNO" in header else 0
    desc_col = None
    for i, h in enumerate(header):
        if h and "DESCRIPTION" in str(h).upper():
            desc_col = i
            break
    if desc_col is None:
        desc_col = 1  # fallback: second column

    gt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sl, desc = row[0], row[desc_col]
        if sl is not None and desc:
            gt[int(sl)] = str(desc).strip()

    sl_values = list(range(50, 0, -1))  # row index 0 (top) = SL 50, etc.

    X, y = [], []
    skipped = []
    for i, sl_val in enumerate(sl_values):
        text = gt.get(sl_val)
        if not text:
            skipped.append((sl_val, "no ground truth")); continue
        known_words = text.split()
        y0, y1 = ROW_BOUNDS[i] + 1, ROW_BOUNDS[i + 1] - 1
        crop = arr[y0:y1, DESC_X0:DESC_X1]
        bw = (crop < INK_THRESHOLD).astype(np.uint8)
        word_spans = split_into_words(bw)

        if len(word_spans) != len(known_words):
            skipped.append((sl_val, f"word count mismatch: "
                             f"found {len(word_spans)} vs expected "
                             f"{len(known_words)} ({known_words})"))
            continue

        row_ok = True
        row_glyphs = []
        for (x0, x1), word in zip(word_spans, known_words):
            word_bw = bw[:, x0:x1]
            tc = tight_crop(word_bw)
            if tc is None:
                row_ok = False; break
            glyphs = slice_word_into_chars(tc, len(word))
            if glyphs is None:
                row_ok = False; break
            row_glyphs.append((word, glyphs))

        if not row_ok:
            skipped.append((sl_val, "char slicing failed"))
            continue

        for word, glyphs in row_glyphs:
            for ch, g in zip(word, glyphs):
                if not ch.isalnum() and ch not in ".:-,()/":
                    continue
                X.append(resize_glyph(g))
                y.append(ch)

    print(f"Harvested {len(X)} labeled letter/symbol glyphs "
          f"from {len(sl_values) - len(skipped)}/{len(sl_values)} rows")
    print(f"Skipped {len(skipped)} rows:")
    for sl, reason in skipped:
        print(f"  SL {sl}: {reason}")

    from collections import Counter
    print("Class distribution:", Counter(y))

    np.save("letter_X.npy", np.array(X))
    np.save("letter_y.npy", np.array(y))


if __name__ == "__main__":
    main()
